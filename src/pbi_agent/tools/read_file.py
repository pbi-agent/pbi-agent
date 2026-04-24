from __future__ import annotations

import csv
import io
import math
import numbers
from zipfile import BadZipFile
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from pbi_agent.tools.output import bound_output
from pbi_agent.tools.types import ToolContext, ToolSpec
from pbi_agent.tools.workspace_access import DEFAULT_MAX_LINES
from pbi_agent.tools.workspace_access import normalize_positive_int
from pbi_agent.tools.workspace_access import open_text_file
from pbi_agent.tools.workspace_access import relative_workspace_path
from pbi_agent.tools.workspace_access import resolve_safe_path

MAX_READ_FILE_OUTPUT_CHARS = 12_000
DATAFRAME_PREVIEW_ROWS = 3
MAX_TABULAR_SCHEMA_CHARS = 8_000
MAX_TABULAR_PREVIEW_CHARS = 2_500
TABULAR_SAMPLE_ROWS = 5_000
MAX_CSV_SNIFF_CHARS = 65_536

_TABULAR_EXTENSIONS = {
    ".csv",
    ".tsv",
    ".xlsx",
    ".xls",
    ".parquet",
    ".feather",
    ".ipc",
    ".arrow",
}
_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp"}

SPEC = ToolSpec(
    name="read_file",
    description=(
        "Read a workspace file. Supports text (with line ranges), "
        "tabular data (CSV/Excel/Parquet), PDF, and DOCX."
    ),
    parameters_schema={
        "type": "object",
        "properties": {
            "path": {
                "type": "string",
                "description": (
                    "File path relative to the workspace root "
                    "(or absolute within workspace)."
                ),
            },
            "start_line": {
                "type": "integer",
                "description": "1-based starting line number. Defaults to 1.",
            },
            "max_lines": {
                "type": "integer",
                "description": "Maximum number of lines to return. Defaults to 200.",
            },
            "encoding": {
                "type": "string",
                "description": "Text encoding to use. Defaults to 'auto'.",
                "default": "auto",
            },
        },
        "required": ["path"],
        "additionalProperties": False,
    },
)


def handle(arguments: dict[str, Any], context: ToolContext) -> dict[str, Any]:
    del context
    path_value = arguments.get("path", "")
    if not isinstance(path_value, str) or not path_value.strip():
        return {"error": "'path' must be a non-empty string."}

    root = Path.cwd().resolve()
    raw_start_line = arguments.get("start_line")
    start_line = (
        raw_start_line if isinstance(raw_start_line, int) and raw_start_line >= 1 else 1
    )
    max_lines = normalize_positive_int(
        arguments.get("max_lines"), default=DEFAULT_MAX_LINES
    )
    encoding = arguments.get("encoding", "auto")

    try:
        target_path = resolve_safe_path(root, path_value)
        if not target_path.exists():
            return {"error": f"path not found: {target_path}"}
        if not target_path.is_file():
            return {"error": f"path is not a file: {target_path}"}

        suffix = target_path.suffix.lower()
        if suffix in _IMAGE_EXTENSIONS:
            return {
                "error": (
                    f"image file is not supported by read_file: {target_path.name}; "
                    "use read_image instead"
                )
            }
        if suffix in {".xlsx", ".xls"}:
            return _handle_excel_workbook(root, target_path)
        if suffix in _TABULAR_EXTENSIONS:
            return _handle_tabular_file(root, target_path, suffix)
        if suffix == ".docx":
            return _handle_docx_file(root, target_path)
        if suffix == ".pdf":
            return _handle_pdf_file(root, target_path)

        selected_lines: list[str] = []
        line_count = 0
        with open_text_file(target_path, encoding=str(encoding)) as text_handle:
            for line_count, line in enumerate(text_handle, start=1):
                if line_count < start_line:
                    continue
                if len(selected_lines) < max_lines:
                    selected_lines.append(line)

        selected = "".join(selected_lines)
        bounded_content, content_truncated = bound_output(
            selected, limit=MAX_READ_FILE_OUTPUT_CHARS
        )
        returned_start_line = start_line if selected_lines else 0
        returned_end_line = (
            returned_start_line + len(selected_lines) - 1 if selected_lines else 0
        )
        has_more_lines = returned_end_line < line_count if returned_end_line else False

        result: dict[str, Any] = {
            "path": relative_workspace_path(root, target_path),
            "start_line": returned_start_line,
            "end_line": returned_end_line,
            "total_lines": line_count,
            "content": bounded_content,
            "has_more_lines": has_more_lines,
        }
        if line_count == 0:
            result["empty"] = True
        if start_line > 1 or has_more_lines:
            result["windowed"] = True
        if content_truncated:
            result["content_truncated"] = True
        return result
    except Exception as exc:
        return {"error": bound_output(str(exc))[0]}


def _handle_tabular_file(root: Path, target_path: Path, suffix: str) -> dict[str, Any]:
    separator: str | None = None
    if suffix in {".csv", ".tsv"}:
        separator = _detect_delimited_separator(target_path, suffix=suffix)
        raw_bytes = target_path.read_bytes()
        row_count = _count_delimited_rows(target_path, separator=separator)
        if row_count > TABULAR_SAMPLE_ROWS:
            dataframe = _read_delimited_dataframe_from_bytes(
                raw_bytes,
                separator=separator,
                nrows=TABULAR_SAMPLE_ROWS,
            )
            result = _summarize_dataframe(
                dataframe,
                actual_rows=row_count,
                sampled_rows=len(dataframe),
            )
            result["path"] = relative_workspace_path(root, target_path)
            return result

    dataframe = _read_tabular_dataframe(
        target_path,
        suffix,
        separator=separator,
    )
    result = _summarize_dataframe(dataframe)
    result["path"] = relative_workspace_path(root, target_path)
    return result


def _handle_excel_workbook(root: Path, target_path: Path) -> dict[str, Any]:
    sheet_metadata = _read_excel_sheet_metadata(target_path)
    if sheet_metadata and any(
        meta["rows"] > TABULAR_SAMPLE_ROWS for meta in sheet_metadata
    ):
        sheets: list[dict[str, Any]] = []
        for meta in sheet_metadata:
            sample_rows = (
                TABULAR_SAMPLE_ROWS if meta["rows"] > TABULAR_SAMPLE_ROWS else None
            )
            dataframe = _read_excel_sheet(
                target_path,
                sheet_name=meta["name"],
                nrows=sample_rows,
            )
            sheet_result = _summarize_dataframe(
                dataframe,
                actual_rows=meta["rows"],
                actual_columns=meta["columns"],
                sampled_rows=len(dataframe) if sample_rows is not None else None,
            )
            sheet_result["name"] = meta["name"]
            sheets.append(sheet_result)

        return {
            "path": relative_workspace_path(root, target_path),
            "sheet_count": len(sheets),
            "sheets": sheets,
        }

    workbook = _read_excel_workbook(target_path)
    sheets: list[dict[str, Any]] = []

    for sheet_name, dataframe in workbook.items():
        sheet_result = _summarize_dataframe(dataframe)
        sheet_result["name"] = sheet_name
        sheets.append(sheet_result)

    return {
        "path": relative_workspace_path(root, target_path),
        "sheet_count": len(sheets),
        "sheets": sheets,
    }


def _handle_docx_file(root: Path, target_path: Path) -> dict[str, Any]:
    full_text = _extract_docx_text(target_path)

    result: dict[str, Any] = {
        "path": relative_workspace_path(root, target_path),
        "content": full_text,
    }
    return result


def _extract_docx_text(path: Path) -> str:
    from docx import Document
    from docx.opc.exceptions import PackageNotFoundError

    try:
        document = Document(str(path))
    except (BadZipFile, PackageNotFoundError) as exc:
        raise ValueError(
            f"unable to read docx file: {path.name} is unreadable or corrupt"
        ) from exc

    parts: list[str] = []

    for block in document.iter_inner_content():
        from docx.table import Table
        from docx.text.paragraph import Paragraph

        if isinstance(block, Paragraph):
            parts.append(block.text)
        elif isinstance(block, Table):
            for row in block.rows:
                parts.append("\t".join(cell.text for cell in row.cells))

    for section in document.sections:
        for header_footer in (section.header, section.footer):
            if header_footer.is_linked_to_previous:
                continue
            for paragraph in header_footer.paragraphs:
                text = paragraph.text.strip()
                if text:
                    parts.append(text)

    return "\n".join(parts)


def _summarize_dataframe(
    dataframe: Any,
    *,
    actual_rows: int | None = None,
    actual_columns: int | None = None,
    sampled_rows: int | None = None,
) -> dict[str, Any]:
    schema = {
        column_name: _display_dtype(dataframe[column_name])
        for column_name in dataframe.columns
    }
    rows, columns = dataframe.shape
    reported_rows = rows if actual_rows is None else actual_rows
    reported_columns = columns if actual_columns is None else actual_columns
    null_counts = {
        column_name: int(dataframe[column_name].isna().sum())
        for column_name in dataframe.columns
    }
    kind_by_column = {
        column_name: _column_kind(dataframe[column_name])
        for column_name in dataframe.columns
    }

    schema_lines = [
        _summarize_tabular_column(
            dataframe[column_name],
            column_name,
            schema[column_name],
            kind=kind_by_column[column_name],
            null_count=null_counts[column_name],
            row_count=rows,
        )
        for column_name in dataframe.columns
    ]
    schema_text, schema_truncated = bound_output(
        "\n".join(f"- {line}" for line in schema_lines),
        limit=MAX_TABULAR_SCHEMA_CHARS,
    )
    preview_text, preview_truncated = bound_output(
        _render_preview_csv(dataframe),
        limit=MAX_TABULAR_PREVIEW_CHARS,
    )

    result: dict[str, Any] = {
        "shape": {
            "rows": reported_rows,
            "columns": reported_columns,
        },
        "summary": _tabular_dataset_summary(
            rows=reported_rows,
            columns=reported_columns,
            kind_by_column=kind_by_column,
            null_counts=null_counts,
            sampled_rows=sampled_rows,
        ),
        "schema": schema_text,
        "preview": preview_text,
    }
    if sampled_rows is not None and sampled_rows < reported_rows:
        result["sampled"] = True
        result["sample_rows"] = sampled_rows
    if schema_truncated:
        result["schema_truncated"] = True
    if preview_truncated:
        result["preview_truncated"] = True
    return result


def _read_tabular_dataframe(
    path: Path, suffix: str, *, separator: str | None = None
) -> Any:
    import pandas as pd
    import pyarrow as pa
    import pyarrow.ipc as ipc

    if suffix == ".csv":
        return _read_delimited_dataframe(path, separator=separator or ",")
    if suffix == ".tsv":
        return _read_delimited_dataframe(path, separator=separator or "\t")
    if suffix == ".parquet":
        return _coerce_temporal_columns(pd.read_parquet(path))
    if suffix == ".feather":
        return _coerce_temporal_columns(pd.read_feather(path))
    if suffix in {".ipc", ".arrow"}:
        with pa.memory_map(str(path), "r") as source:
            try:
                dataframe = ipc.RecordBatchFileReader(source).read_all().to_pandas()
            except pa.ArrowInvalid:
                source.seek(0)
                dataframe = ipc.RecordBatchStreamReader(source).read_all().to_pandas()
        return _coerce_temporal_columns(dataframe)
    raise ValueError(f"unsupported tabular file type: {suffix}")


def _read_excel_workbook(path: Path) -> dict[str, Any]:
    import pandas as pd

    workbook = pd.read_excel(path, sheet_name=None)
    return {
        str(sheet_name): _coerce_temporal_columns(dataframe)
        for sheet_name, dataframe in workbook.items()
    }


def _read_excel_sheet(path: Path, *, sheet_name: str, nrows: int | None) -> Any:
    import pandas as pd

    dataframe = pd.read_excel(path, sheet_name=sheet_name, nrows=nrows)
    return _coerce_temporal_columns(dataframe)


def _read_delimited_dataframe(path: Path, *, separator: str) -> Any:
    raw_bytes = path.read_bytes()
    return _read_delimited_dataframe_from_bytes(raw_bytes, separator=separator)


def _read_delimited_dataframe_from_bytes(
    raw_bytes: bytes,
    *,
    separator: str,
    nrows: int | None = None,
) -> Any:
    import pandas as pd

    normalized_bytes = _normalize_delimited_bytes(raw_bytes)
    dataframe = pd.read_csv(
        io.BytesIO(normalized_bytes),
        sep=separator,
        low_memory=False,
        nrows=nrows,
    )
    return _coerce_temporal_columns(dataframe)


def _normalize_delimited_bytes(raw_bytes: bytes) -> bytes:
    if b"\r" in raw_bytes and b"\n" not in raw_bytes:
        return raw_bytes.replace(b"\r", b"\n")
    return raw_bytes


def _detect_delimited_separator(path: Path, *, suffix: str) -> str:
    if suffix == ".tsv":
        return "\t"
    if suffix != ".csv":
        raise ValueError(f"unsupported delimited file type: {suffix}")

    with open_text_file(path, encoding="auto") as text_handle:
        sample = text_handle.read(MAX_CSV_SNIFF_CHARS)

    if not sample:
        return ","

    normalized_sample = sample.replace("\r", "\n") if "\n" not in sample else sample
    try:
        dialect = csv.Sniffer().sniff(normalized_sample, delimiters=",;\t|")
    except csv.Error:
        return ","
    return str(dialect.delimiter)


def _count_delimited_rows(path: Path, *, separator: str) -> int:
    with open_text_file(path, encoding="auto") as text_handle:
        reader = csv.reader(text_handle, delimiter=separator)
        try:
            next(reader)
        except StopIteration:
            return 0
        return sum(1 for _ in reader)


def _read_excel_sheet_metadata(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return []

    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError:
        return []

    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError):
        return []

    try:
        metadata: list[dict[str, Any]] = []
        for worksheet in workbook.worksheets:
            rows, columns = _read_xlsx_sheet_bounds(worksheet)
            metadata.append(
                {
                    "name": str(worksheet.title),
                    "rows": max(rows - 1, 0),
                    "columns": columns,
                }
            )
        return metadata
    finally:
        workbook.close()


def _read_xlsx_sheet_bounds(worksheet: Any) -> tuple[int, int]:
    worksheet.reset_dimensions()
    max_row = 0
    max_column = 0

    for row in worksheet.iter_rows():
        populated_cells = [cell for cell in row if cell.value is not None]
        if not populated_cells:
            continue
        max_row = max(max_row, max(cell.row for cell in populated_cells))
        max_column = max(max_column, max(cell.column for cell in populated_cells))

    return (max_row, max_column)


def _excel_column_letters_to_index(letters: str) -> int:
    index = 0
    for character in letters:
        index = index * 26 + (ord(character) - ord("A") + 1)
    return index


def _coerce_temporal_columns(dataframe: Any) -> Any:
    from pandas.api import types as ptypes

    temporal_formats: dict[str, tuple[str, ...]] = {
        "date": (
            "%Y-%m-%d",
            "%m/%d/%Y",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%dT%H:%M:%S",
        ),
        "datetime": ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"),
    }
    temporal_kinds = (
        "date",
        "datetime",
    )
    result = dataframe.copy()

    for column_name in dataframe.columns:
        series = dataframe[column_name]
        if not (
            ptypes.is_string_dtype(series.dtype) or ptypes.is_object_dtype(series.dtype)
        ):
            continue

        non_null = series.dropna()
        if non_null.empty or not all(
            _is_temporal_candidate(value) for value in non_null
        ):
            continue

        cleaned = series.map(_normalize_temporal_candidate, na_action="ignore").astype(
            "string"
        )
        cleaned = cleaned.str.strip()
        for temporal_kind in temporal_kinds:
            parsed = _parse_temporal_series(cleaned, temporal_formats[temporal_kind])
            if parsed is not None and int(parsed.notna().sum()) == len(non_null):
                if temporal_kind == "date":
                    result[column_name] = parsed.dt.date
                else:
                    result[column_name] = parsed
                break

    return result


def _parse_temporal_series(series: Any, formats: tuple[str, ...]) -> Any:
    import pandas as pd

    parsed = pd.Series(pd.NaT, index=series.index, dtype="datetime64[ns]")
    remaining = series.notna()

    for format_string in formats:
        if not bool(remaining.any()):
            break

        current = pd.to_datetime(
            series.where(remaining),
            format=format_string,
            errors="coerce",
        )
        matched = current.notna() & remaining
        if bool(matched.any()):
            parsed.loc[matched] = current.loc[matched]
            remaining = remaining & ~matched

    if bool(remaining.any()):
        return None
    return parsed


def _is_temporal_candidate(value: Any) -> bool:
    return isinstance(value, (str, date, datetime))


def _normalize_temporal_candidate(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(sep="T")
    if isinstance(value, date):
        return value.isoformat()
    return value


def _is_categorical_column(series: Any) -> bool:
    if _is_categorical_dtype(series):
        return True
    if series.empty or not _is_text_series(series):
        return False

    non_null = series.dropna()
    if non_null.empty:
        return False
    distinct = int(non_null.nunique(dropna=True))
    unique_ratio = distinct / len(non_null)
    return unique_ratio <= 0.2 and distinct <= 100


def _column_kind(series: Any) -> str:
    from pandas.api import types as ptypes

    if ptypes.is_numeric_dtype(series.dtype) and not _is_boolean_series(series):
        return "numeric"
    if _is_temporal_series(series):
        return "datetime"
    if _is_boolean_series(series):
        return "boolean"
    if _is_categorical_column(series):
        return "categorical"
    if _is_text_series(series) or _is_categorical_dtype(series):
        return "text"
    return "other"


def _display_dtype(series: Any) -> str:
    from pandas.api import types as ptypes

    if _is_boolean_series(series):
        return "Boolean"
    if _is_date_only_series(series):
        return "Date"
    if _is_temporal_series(series):
        return "Datetime"
    if ptypes.is_integer_dtype(series.dtype):
        return "Int64"
    if ptypes.is_float_dtype(series.dtype):
        return "Float64"
    if _is_categorical_dtype(series):
        return "Categorical"
    if _is_text_series(series):
        return "String"
    return str(series.dtype)


def _tabular_dataset_summary(
    *,
    rows: int,
    columns: int,
    kind_by_column: dict[str, str],
    null_counts: dict[str, int],
    sampled_rows: int | None = None,
) -> str:
    kind_order = ["numeric", "datetime", "categorical", "text", "boolean", "other"]
    kind_counts = {
        kind: sum(1 for value in kind_by_column.values() if value == kind)
        for kind in kind_order
    }
    mix_parts = [f"{count} {kind}" for kind, count in kind_counts.items() if count > 0]
    total_nulls = sum(null_counts.values())
    missing_summary = (
        "no missing values"
        if total_nulls == 0
        else f"{total_nulls} missing values across {sum(1 for count in null_counts.values() if count > 0)} columns"
    )

    parts = [f"{rows} rows x {columns} columns"]
    if sampled_rows is not None and sampled_rows < rows:
        parts.append(f"schema/stats sampled from first {sampled_rows} rows")
    if mix_parts:
        parts.append("column mix: " + ", ".join(mix_parts))
    if sampled_rows is not None and sampled_rows < rows:
        parts.append(f"missing values in sample: {missing_summary}")
    else:
        parts.append(f"missing values: {missing_summary}")
    return "; ".join(parts) + "."


def _summarize_tabular_column(
    series: Any,
    column_name: str,
    dtype: Any,
    *,
    kind: str,
    null_count: int,
    row_count: int,
) -> str:
    parts = [f"{column_name}: {dtype}"]
    if kind != "other":
        parts.append(f"kind={kind}")
    if null_count:
        null_ratio = (null_count / row_count) * 100 if row_count else 0.0
        parts.append(f"nulls={null_count} ({_format_number(null_ratio)}%)")

    non_null = series.dropna()
    if non_null.empty:
        parts.append("all values null")
        return "; ".join(parts)

    if kind == "numeric":
        parts.extend(
            [
                f"min={_format_scalar(non_null.min())}",
                f"max={_format_scalar(non_null.max())}",
                f"mean={_format_scalar(non_null.mean())}",
            ]
        )
        return "; ".join(parts)

    if kind == "datetime":
        parts.append(
            f"range={_format_scalar(non_null.min())}..{_format_scalar(non_null.max())}"
        )
        return "; ".join(parts)

    if kind == "boolean":
        parts.append(
            "values=" + ", ".join(_ordered_examples(non_null.to_list(), limit=2))
        )
        return "; ".join(parts)

    if kind == "categorical":
        distinct = int(non_null.nunique(dropna=True))
        examples = _top_examples(non_null, limit=5)
        parts.append(f"distinct={distinct}")
        if examples:
            parts.append("examples=" + ", ".join(examples))
        return "; ".join(parts)

    if kind == "text":
        examples = _ordered_examples(non_null.to_list(), limit=3)
        if examples:
            parts.append("examples=" + ", ".join(examples))
        return "; ".join(parts)

    return "; ".join(parts)


def _render_preview_csv(dataframe: Any) -> str:
    preview = dataframe.head(DATAFRAME_PREVIEW_ROWS)
    if preview.empty:
        return ""

    buffer = io.StringIO()
    writer = csv.DictWriter(
        buffer,
        fieldnames=preview.columns,
        lineterminator="\n",
    )
    writer.writeheader()
    for row in preview.to_dict(orient="records"):
        writer.writerow(
            {
                column_name: _format_preview_value(row.get(column_name))
                for column_name in preview.columns
            }
        )
    return buffer.getvalue()


def _top_examples(series: Any, *, limit: int) -> list[str]:
    values = series.value_counts(dropna=True).index.tolist()
    return _ordered_examples(values, limit=limit)


def _is_text_series(series: Any) -> bool:
    from pandas.api import types as ptypes

    if ptypes.is_string_dtype(series.dtype):
        return True
    if not ptypes.is_object_dtype(series.dtype):
        return False

    non_null = series.dropna()
    return non_null.empty or all(isinstance(value, str) for value in non_null)


def _is_categorical_dtype(series: Any) -> bool:
    import pandas as pd

    return isinstance(series.dtype, pd.CategoricalDtype)


def _is_boolean_series(series: Any) -> bool:
    from pandas.api import types as ptypes

    if ptypes.is_bool_dtype(series.dtype):
        return True
    if not ptypes.is_object_dtype(series.dtype):
        return False

    non_null = series.dropna()
    return not non_null.empty and all(isinstance(value, bool) for value in non_null)


def _is_date_only_series(series: Any) -> bool:
    from pandas.api import types as ptypes

    non_null = series.dropna()
    if non_null.empty:
        return False
    if _is_categorical_dtype(series):
        return False
    if all(
        isinstance(value, date) and not isinstance(value, datetime)
        for value in non_null
    ):
        return True
    if ptypes.is_datetime64_any_dtype(series.dtype):
        timestamps = non_null.dt
        return bool(
            (
                (timestamps.hour == 0)
                & (timestamps.minute == 0)
                & (timestamps.second == 0)
                & (timestamps.microsecond == 0)
                & (timestamps.nanosecond == 0)
            ).all()
        )
    if not all(isinstance(value, datetime) for value in non_null):
        return False
    return bool(
        all(
            value.hour == 0
            and value.minute == 0
            and value.second == 0
            and value.microsecond == 0
            for value in non_null
        )
    )


def _is_temporal_series(series: Any) -> bool:
    from pandas.api import types as ptypes

    if ptypes.is_datetime64_any_dtype(series.dtype) or ptypes.is_timedelta64_dtype(
        series.dtype
    ):
        return True
    if _is_date_only_series(series):
        return True
    if ptypes.is_object_dtype(series.dtype):
        non_null = series.dropna()
        return not non_null.empty and all(isinstance(value, time) for value in non_null)
    return False


def _ordered_examples(values: list[Any], *, limit: int) -> list[str]:
    examples: list[str] = []
    seen: set[str] = set()
    for value in values:
        rendered = _format_scalar(value)
        if rendered in seen:
            continue
        seen.add(rendered)
        examples.append(rendered)
        if len(examples) >= limit:
            break
    return examples


def _format_preview_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return str(value)
    return _format_scalar(value)


def _format_scalar(value: Any) -> str:
    import pandas as pd

    if pd.isna(value):
        return "null"
    if value is None:
        return "null"
    if isinstance(value, datetime):
        if (
            value.hour == 0
            and value.minute == 0
            and value.second == 0
            and value.microsecond == 0
        ):
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, bool):
        return str(value).lower()
    if isinstance(value, numbers.Integral):
        return str(value)
    if isinstance(value, numbers.Real):
        return _format_number(float(value))
    return str(value)


def _format_number(value: float | numbers.Real) -> str:
    value = float(value)
    if math.isnan(value):
        return "null"
    if value.is_integer():
        return str(int(value))
    return f"{value:.4f}".rstrip("0").rstrip(".")


def _handle_pdf_file(root: Path, target_path: Path) -> dict[str, Any]:
    from pypdf import PdfReader

    reader = PdfReader(str(target_path))
    page_text = [page.extract_text() or "" for page in reader.pages]
    full_text = "\n".join(page_text)

    metadata: dict[str, Any] = {"pages": len(reader.pages)}
    if reader.metadata:
        if reader.metadata.title:
            metadata["title"] = reader.metadata.title
        if reader.metadata.author:
            metadata["author"] = reader.metadata.author

    result: dict[str, Any] = {
        "path": relative_workspace_path(root, target_path),
        "content": full_text,
        "metadata": metadata,
    }
    return result
